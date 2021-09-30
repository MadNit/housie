"""
__author__ = 'Vijai'
"""
import os
from pathlib import Path
import random
from itertools import permutations
import pprint
from openpyxl import Workbook, load_workbook

house_xl_tmpl = os.path.join("C:\\", "Users", "vijai", "Downloads", "Housie", "tmpl_macro.xlsm")
ZEROCOM = [0, 0, 0, 0, 1, 1, 1, 1, 1]
ALLCOM = set()
for ival in permutations(ZEROCOM):
    ALLCOM.add(ival)
LENCOM = len(ALLCOM)
ALLCOM = sorted(list(ALLCOM))
for ind in range(LENCOM):
    ALLCOM[ind] = list(ALLCOM[ind])


def generate_tickets(num_tickets=1):
    """
    Generates number of tickets
    """
    tkts_l = []
    for numt in range(num_tickets):
        tkts_l.append(__gen_ticket())
    return tkts_l


def __gen_ticket():
    """
    Generates one ticket

    num_tickets: number of tickets to generate
    """
    col1 = sorted(random.sample(range(1, 9), 3))
    col2 = sorted(random.sample(range(10, 19), 3))
    col3 = sorted(random.sample(range(20, 29), 3))
    col4 = sorted(random.sample(range(30, 39), 3))
    col5 = sorted(random.sample(range(40, 49), 3))
    col6 = sorted(random.sample(range(50, 59), 3))
    col7 = sorted(random.sample(range(60, 69), 3))
    col8 = sorted(random.sample(range(70, 79), 3))
    col9 = sorted(random.sample(range(80, 90), 3))

    tline1 = [col1[0], col2[0], col3[0], col4[0], col5[0], col6[0], col7[0], col8[0], col9[0]]
    tline2 = [col1[1], col2[1], col3[1], col4[1], col5[1], col6[1], col7[1], col8[1], col9[1]]
    tline3 = [col1[2], col2[2], col3[2], col4[2], col5[2], col6[2], col7[2], col8[2], col9[2]]

    tkt_sam = None
    while True:
        tkt_sam = random.sample(range(1, LENCOM), 3)
        try:
            if ALLCOM[tkt_sam[0]][0] + ALLCOM[tkt_sam[1]][0] + ALLCOM[tkt_sam[2]][0] < 1 or \
                    ALLCOM[tkt_sam[0]][1] + ALLCOM[tkt_sam[1]][1] + ALLCOM[tkt_sam[2]][1] < 1 or \
                    ALLCOM[tkt_sam[0]][2] + ALLCOM[tkt_sam[1]][2] + ALLCOM[tkt_sam[2]][2] < 1 or \
                    ALLCOM[tkt_sam[0]][3] + ALLCOM[tkt_sam[1]][3] + ALLCOM[tkt_sam[2]][3] < 1 or \
                    ALLCOM[tkt_sam[0]][4] + ALLCOM[tkt_sam[1]][4] + ALLCOM[tkt_sam[2]][4] < 1 or \
                    ALLCOM[tkt_sam[0]][5] + ALLCOM[tkt_sam[1]][5] + ALLCOM[tkt_sam[2]][5] < 1 or \
                    ALLCOM[tkt_sam[0]][6] + ALLCOM[tkt_sam[1]][6] + ALLCOM[tkt_sam[2]][6] < 1 or \
                    ALLCOM[tkt_sam[0]][7] + ALLCOM[tkt_sam[1]][7] + ALLCOM[tkt_sam[2]][7] < 1 or \
                    ALLCOM[tkt_sam[0]][8] + ALLCOM[tkt_sam[1]][8] + ALLCOM[tkt_sam[2]][8] < 1:
                continue
        except Exception as e:
            print(e)
        break

    try:
        tline1_0 = ALLCOM[tkt_sam[0]]
        tline2_0 = ALLCOM[tkt_sam[1]]
        tline3_0 = ALLCOM[tkt_sam[2]]
    except Exception as e1:
        print("error as: ", e1)
        raise Exception('Error here...........')

    for val in range(9):
        val1 = tline1[val]
        tline1[val] = tline1_0[val] * val1

        val1 = tline2[val]
        tline2[val] = tline2_0[val] * val1

        val1 = tline3[val]
        tline3[val] = tline3_0[val] * val1

    return tline1, tline2, tline3


def main():
    tkt = generate_tickets(24)
    pp = pprint.PrettyPrinter(indent=4)

    new_tickets = "housie_tickets_"
    npath = Path(house_xl_tmpl)
    new_housie_tkt_xl = os.path.join(npath.parent, new_tickets + ".xlsm")
    # Load housie excel template
    wb = load_workbook(house_xl_tmpl, read_only=False, keep_vba=True)
    sh = wb['HOUSIE']

    rowcol = [(9, 1), (13, 1), (18, 1), (22, 1), (27, 1), (31, 1),
              (9, 11), (13, 11), (18, 11), (22, 11), (27, 11), (31, 11),
              (9, 21), (13, 21), (18, 21), (22, 21), (27, 21), (31, 21),
              (9, 31), (13, 31), (18, 31), (22, 31), (27, 31), (31, 31),
              (9, 41), (13, 41), (18, 41), (22, 41), (27, 41), (31, 41)
              ]

    for ct, mtkt in enumerate(tkt):
        row = rowcol[ct][0]
        for v in mtkt:
            col = rowcol[ct][1]
            for cv in v:
                cl = sh.cell(row, col)
                if cv == 0:
                    cl.value = ''
                else:
                    cl.value = cv
                col += 1
            row += 1
    cl = sh.cell(5, 5)
    cl.value = 1

    # Update sheet NOS
    sh1 = wb['NOS']
    # Get all the random numbers
    l1 = random.sample(range(1, 91), 90)

    for counter, val in enumerate(l1):
        cl = sh1.cell(counter+1, 1)
        cl.value = val
        # print(val)

    sh1.column_dimensions.group('A', hidden=True)
    sh1.row_dimensions.group(1, 100, hidden=True)
    sh1.protection.sheet = True
    wb.save(new_housie_tkt_xl)

    # for val in tkt:
    #     pp.pprint(val)


if __name__ == "__main__":
    main()




